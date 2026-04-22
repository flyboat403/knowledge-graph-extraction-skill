#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文档知识图谱结构化抽取脚本

用法:
    # 方式1: 从 LLM JSON 输出生成 Excel（推荐）
    python extract_knowledge_graph.py --json knowledge_nodes.json --template template.xlsx --output output.xlsx

    # 方式2: 提取文档内容供 LLM 处理
    python extract_knowledge_graph.py --source document.docx --extract-text --output document_content.txt

    # 方式3: 仅解析模板
    python extract_knowledge_graph.py --template template.xlsx --dry-run

技术架构:
    文档 → 文本提取(--source --extract-text) → LLM处理 → JSON输出(--json) → Excel生成
    
    正则匹配已被移除，所有语义理解由 LLM 完成。

功能:
    1. 解析 xlsx 模板 A1 单元格的格式规则
    2. 提取 docx/pdf 文档内容供 LLM 处理
    3. 解析 LLM 输出的 JSON 格式知识节点
    4. 自动生成后置关系（前置关系的反向）
    5. 验证格式约束（包括节点类型与知识点分类的约束）
    6. 输出符合导入规范的 CSV/Excel 文件

关键输出：A列(节点类型) + B-H列(层级) + I-K列(三类关系) + 其他属性列
"""

import argparse
import csv
import json
import sys
import io
import os

# 设置控制台输出编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

from dataclasses import dataclass, field
from typing import Dict, List

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ==================== 常量定义 ====================

# 节点类型常量
NODE_TYPES = ["分类", "知识点"]

# 知识点分类（仅知识点节点可填写）
VALID_CATEGORIES = ["事实性", "概念性", "程序性", "元认知"]

# 标签
VALID_TAGS = ["重点", "难点", "考点", "课程思政"]

# 表头（15列）
EXPECTED_HEADERS = [
    '节点类型', '节点名称', '节点名称', '节点名称', '节点名称',
    '节点名称', '节点名称', '节点名称',
    '前置节点', '后置节点', '关联节点',
    '标签', '知识点分类', '节点说明', '教学目标'
]

# 层级名称
LEVEL_NAMES = {
    1: '一级(课程)', 
    2: '二级(模块)', 
    3: '三级(章节)', 
    4: '四级(主题)', 
    5: '五级(知识点)', 
    6: '六级', 
    7: '七级'
}

# 层级到列映射：从B列开始（B=一级, H=七级）
# 新模板结构：A列=节点类型, B-H列=层级节点名称
LEVEL_TO_COL = {1: "B", 2: "C", 3: "D", 4: "E", 5: "F", 6: "G", 7: "H"}

# ==================== 数据结构定义 ====================

@dataclass
class KnowledgeNode:
    """知识点节点"""
    name: str
    level: int  # 1=B(一级), 2=C(二级), ..., 7=H(七级)
    node_type: str = "知识点"  # 新增：节点类型，"分类" 或 "知识点"
    tags: str = ""
    category: str = ""  # 知识点分类（仅知识点节点时填写）
    objective: str = ""
    description: str = ""
    pre_requisites: str = ""  # I列 - 前置节点
    post_requisites: str = ""  # J列 - 后置节点  
    related: str = ""          # K列 - 关联节点
    children: list = field(default_factory=list)

# ==================== 节点类型判断 ====================

def determine_node_type(level: int, name: str = "") -> str:
    """根据层级确定节点类型
    
    规则：
    - Level 1-3 默认为"分类"（课程、模块、章节等结构容器）
    - Level 4-7 默认为"知识点"（主题、具体知识点等学习内容）
    """
    if level <= 3:
        return "分类"
    return "知识点"

# ==================== 模板解析 ====================

def parse_template(filepath: str) -> dict:
    """解析 xlsx 模板，返回格式规则"""
    wb = load_workbook(filepath)
    ws = wb.active
    
    a1_value = ws['A1'].value or ""
    
    headers = []
    if ws.max_row > 1:
        headers = [ws.cell(row=2, column=i).value for i in range(1, min(16, ws.max_column + 1))]
    
    return {
        "a1_rules": a1_value,
        "headers": headers,
        "max_column": ws.max_column,
        "sheet_name": ws.title
    }

# ==================== 文档内容提取 ====================

def extract_text_from_docx(filepath: str) -> str:
    """从 docx 文档提取纯文本"""
    try:
        from docx import Document
        doc = Document(filepath)
        paragraphs = []
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                paragraphs.append(text)
        return "\n".join(paragraphs)
    except ImportError:
        print("Error: python-docx not installed. Run: pip install python-docx", file=sys.stderr)
        return ""
    except Exception as e:
        print(f"Error parsing docx: {e}", file=sys.stderr)
        return ""

def extract_text_from_pdf(filepath: str) -> str:
    """从 PDF 文档提取纯文本"""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if text.strip():
                    text_parts.append(text.strip())
        return "\n\n".join(text_parts)
    except ImportError:
        print("Error: pdfplumber not installed. Run: pip install pdfplumber", file=sys.stderr)
        return ""
    except Exception as e:
        print(f"Error parsing PDF: {e}", file=sys.stderr)
        return ""

def extract_document_content(source_path: str) -> str:
    """提取文档内容"""
    if source_path.endswith('.docx'):
        return extract_text_from_docx(source_path)
    elif source_path.endswith('.pdf'):
        return extract_text_from_pdf(source_path)
    else:
        print(f"Error: 不支持的文档格式: {source_path}", file=sys.stderr)
        print("支持的格式: .docx, .pdf", file=sys.stderr)
        return ""

# ==================== JSON 解析（LLM 输出） ====================

def parse_llm_json(filepath: str) -> List[KnowledgeNode]:
    """解析 LLM 输出的 JSON 文件，转换为 KnowledgeNode 列表"""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    if not isinstance(data, list):
        raise ValueError("JSON 根元素必须是数组")
    
    nodes = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            print(f"警告: 第 {i+1} 个元素不是对象，跳过", file=sys.stderr)
            continue
        
        name = item.get('name', '')
        if not name:
            print(f"警告: 第 {i+1} 个知识点缺少 name 字段，跳过", file=sys.stderr)
            continue
        
        level = item.get('level', 5)
        
        # 获取节点类型，如果未指定则根据层级自动判断
        node_type = item.get('node_type', '')
        if not node_type:
            node_type = determine_node_type(level, name)
        
        # 处理前置和关联关系（支持数组或字符串）
        pre_req = item.get('pre_requisites', [])
        if isinstance(pre_req, list):
            pre_req = ';'.join(str(p).strip() for p in pre_req if p)
        else:
            pre_req = str(pre_req).strip() if pre_req else ''
        
        related = item.get('related', [])
        if isinstance(related, list):
            related = ';'.join(str(r).strip() for r in related if r)
        else:
            related = str(related).strip() if related else ''
        
        # 知识点分类：仅知识点节点才填写
        category = item.get('category', '')
        if node_type == "分类":
            category = ""  # 分类节点必须留空
        
        node = KnowledgeNode(
            name=name,
            level=level,
            node_type=node_type,
            tags=item.get('tags', ''),
            category=category,
            objective=item.get('objective', ''),
            description=item.get('description', ''),
            pre_requisites=pre_req,
            related=related,
        )
        nodes.append(node)
    
    return nodes

# ==================== 关系生成 ====================

def generate_post_relations(nodes: List[KnowledgeNode]) -> None:
    """生成后置关系（前置关系的反向）"""
    name_to_node: Dict[str, KnowledgeNode] = {n.name: n for n in nodes}
    post_relations: Dict[str, List[str]] = {}
    
    for node in nodes:
        if node.pre_requisites:
            pre_list = [p.strip() for p in node.pre_requisites.split(';') if p.strip()]
            for pre_name in pre_list:
                if pre_name not in post_relations:
                    post_relations[pre_name] = []
                if node.name not in post_relations[pre_name]:
                    post_relations[pre_name].append(node.name)
    
    for node in nodes:
        if node.name in post_relations:
            node.post_requisites = ';'.join(post_relations[node.name])

def validate_relations(nodes: List[KnowledgeNode]) -> List[str]:
    """验证关系约束"""
    errors = []
    all_names = {n.name for n in nodes}
    
    for node in nodes:
        # 检查前置关系
        if node.pre_requisites:
            for pre in node.pre_requisites.split(';'):
                pre = pre.strip()
                if pre and pre not in all_names:
                    errors.append(f"节点 '{node.name}' 的前置知识点 '{pre}' 不存在")
        
        # 检查关联关系
        if node.related:
            for rel in node.related.split(';'):
                rel = rel.strip()
                if rel and rel not in all_names:
                    errors.append(f"节点 '{node.name}' 的关联知识点 '{rel}' 不存在")
    
    return errors

# ==================== 层级树构建 ====================

def build_knowledge_tree(nodes: List[KnowledgeNode]) -> List[KnowledgeNode]:
    """根据层级构建知识树，建立父子关系
    
    要求：输入的节点必须按深度优先顺序排列（由 LLM Prompt 保证）
    
    Layer 2: 即时修正 - 如果parent是知识点，改为分类（知识点不能有子节点）
    """
    level_stack: Dict[int, KnowledgeNode] = {}
    root_nodes = []
    corrections = []  # 记录修正
    
    for node in nodes:
        # 清理栈中层级 >= 当前层级的节点
        for l in list(level_stack.keys()):
            if l >= node.level:
                del level_stack[l]
        
        # 父节点是当前层级-1的栈顶节点
        parent = level_stack.get(node.level - 1)
        
        if parent:
            # Layer 2: 即时修正 - 知识点节点不能有子节点
            # 如果parent是知识点，改为分类
            if parent.node_type == "知识点":
                parent.node_type = "分类"
                parent.category = ""  # 清空知识点分类
                corrections.append({
                    'name': parent.name,
                    'level': parent.level,
                    'children_count': len(parent.children) + 1
                })
            
            parent.children.append(node)
        else:
            root_nodes.append(node)
        
        # 将当前节点压入其层级的栈
        level_stack[node.level] = node
    
    # 输出修正统计（如果有修正）
    if corrections:
        print(f"  ⚠️ 修正节点类型: {len(corrections)} 个（知识点→分类）")
        for c in corrections[:3]:
            print(f"    - '{c['name']}' (Level {c['level']})")
        if len(corrections) > 3:
            print(f"    - ... 还有 {len(corrections) - 3} 个")
    
    return root_nodes

# ==================== 扁平化输出 ====================

def flatten_nodes(nodes: List[KnowledgeNode]) -> List[Dict]:
    """将知识节点列表扁平化为 Excel 行数据
    
    新模板列结构：
    - A列: 节点类型（分类/知识点）
    - B-H列: 节点名称层级（一级=B, 七级=H）
    - I列: 前置节点
    - J列: 后置节点
    - K列: 关联节点
    - L列: 标签
    - M列: 知识点分类（分类节点必须留空）
    - N列: 节点说明
    - O列: 教学目标
    """
    rows = []
    
    for node in nodes:
        # 初始化所有列（A-O）
        row = {col: "" for col in "ABCDEFGHIJKLMNO"}
        
        # A列：节点类型
        row["A"] = node.node_type
        
        # B-H列：层级节点名称（只有一个非空）
        if node.level in LEVEL_TO_COL:
            row[LEVEL_TO_COL[node.level]] = node.name
        
        # I-K列：关系
        row.update({
            "I": node.pre_requisites,
            "J": node.post_requisites, 
            "K": node.related,
            "L": node.tags,
        })
        
        # M列：知识点分类
        # 重要规则：分类节点时M列必须留空
        if node.node_type == "知识点":
            row["M"] = node.category
        else:
            row["M"] = ""  # 分类节点不填写知识分类
        
        # N列：节点说明
        row["N"] = node.description
        
        # O列：教学目标
        row["O"] = node.objective
        
        rows.append(row)
        
        if node.children:
            rows.extend(flatten_nodes(node.children))
    
    return rows

# ==================== Excel/CSV 生成 ====================

def generate_excel(rows: List[Dict], output_path: str, template_rules: str = ""):
    """生成 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "知识点抽取结果"
    
    # A1单元格使用说明（匹配新模板）
    instructions = template_rules or """使用说明：
1. 后面带"*"的为必填项
2. 节点类型：分类、知识点
3. A列仅支持填写一个节点类型，B列至H列每行仅支持填写一个节点
4. I列至K列填写对应节点的前置、后置和关联节点，多个节点之间用英文分号";"隔开
5. 任意两个节点之间仅支持存在一种关系（前置、后置或关联），新导入的关系将覆盖旧关系
6. 固定标签包括：重点、难点、考点、课程思政，可根据需要自定义标签
7. 知识点分类包括：事实性、概念性、程序性、元认知，每个知识点只能填入一个知识点分类，分类节点不支持填写知识点分类
8. 节点说明以及教学目标仅支持输入文本"""
    
    ws['A1'] = instructions
    ws.merge_cells('A1:O1')
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')

    # 新模板表头
    headers = EXPECTED_HEADERS

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 列宽度设置
    col_widths = {
        'A': 12,  # 节点类型
        'B': 25, 'C': 20, 'D': 25, 'E': 25, 'F': 25, 'G': 15, 'H': 15,  # 层级节点
        'I': 35, 'J': 35, 'K': 35,  # 关系
        'L': 15,  # 标签
        'M': 12,  # 知识点分类
        'N': 40,  # 节点说明
        'O': 35   # 教学目标
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 写入数据行
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, key in enumerate(['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O'], 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key) or "")
    
    wb.save(output_path)
    print(f"✅ Excel 文件已保存: {output_path}")

def generate_csv(rows: List[Dict], output_path: str):
    """生成 CSV 文件"""
    headers = EXPECTED_HEADERS

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row_data in rows:
            writer.writerow([row_data.get(c, '') for c in 'ABCDEFGHIJKLMNO'])
    
    print(f"✅ CSV 文件已保存: {output_path}")

# ==================== 验证 ====================

def validate_rows(rows: List[Dict]) -> List[str]:
    """验证行数据是否符合约束"""
    errors = []
    
    # Helper functions for Layer 3 validation
    def find_row_level(row: Dict) -> int:
        """找出行的层级（B-H列中非空的列位置）"""
        for idx, col in enumerate('BCDEFGH', 1):
            if row.get(col, ""):
                return idx
        return 0
    
    def get_row_name(row: Dict, level: int) -> str:
        """获取行在指定层级的节点名称"""
        col = chr(64 + level + 1)  # B=66, level=1 → B
        return row.get(col, "")
    
    for i, row in enumerate(rows, 3):
        # 验证节点类型（A列）
        node_type = row.get('A', '')
        if node_type not in NODE_TYPES:
            errors.append(f"行 {i}: 节点类型无效: {node_type}，应为'分类'或'知识点'")
        
        # 验证层级节点（B-H列只有一个非空）
        level_vals = [row.get(col, "") for col in 'BCDEFGH']
        non_empty_count = sum(1 for val in level_vals if val)
        if non_empty_count != 1:
            errors.append(f"行 {i}: 每行应只有1个节点名称，实际{non_empty_count}个")
        
        # 验证知识点分类（M列）
        # 重要规则：分类节点时M列必须为空
        category = row.get('M', '')
        if node_type == "分类" and category:
            errors.append(f"行 {i}: 分类节点的知识点分类(M列)必须为空")
        if node_type == "知识点" and category and category not in VALID_CATEGORIES:
            errors.append(f"行 {i}: 知识点分类无效: {category}")
        
        # Layer 3: 知识点子节点约束检查（双重保险）
        # 如果当前行是知识点，检查下一行是否是子节点
        if node_type == "知识点":
            row_idx = i - 3  # rows列表索引
            if row_idx < len(rows) - 1:
                next_row = rows[row_idx + 1]
                current_level = find_row_level(row)
                next_level = find_row_level(next_row)
                
                # 如果下一行层级更深，说明当前知识点有子节点
                if next_level > current_level:
                    name = get_row_name(row, current_level)
                    errors.append(
                        f"行 {i}: 知识点 '{name}' 有子节点（修正遗漏？）"
                    )
        
        # 验证分隔符（使用英文分号）
        for col in 'IJK':
            val = row.get(col, '')
            if val and '；' in val:
                errors.append(f"行 {i}: {col}列使用中文分号，应使用英文分号';'")

    return errors

def validate_headers(filepath: str) -> List[str]:
    """验证Excel文件的表头是否符合预期

    读取Excel文件第2行（row 2，索引为1）的表头，对比EXPECTED_HEADERS。
    返回错误列表，错误格式：'列{X}实际为'{actual}',应为'{expected}''
    """
    errors = []
    wb = load_workbook(filepath)
    ws = wb.active

    for col_idx, expected_header in enumerate(EXPECTED_HEADERS, 1):
        actual_header = ws.cell(row=2, column=col_idx).value
        if actual_header != expected_header:
            col_letter = chr(64 + col_idx)
            errors.append(f"列{col_letter}实际为'{actual_header}',应为'{expected_header}'")

    return errors

# ==================== 语义质量校验 ====================

def validate_quality(nodes: List[KnowledgeNode]) -> Dict[str, List[str]]:
    """语义质量校验（供Agent审查和修正）
    
    返回问题字典，每个类别包含具体问题列表。
    Agent应根据此输出逐一修正JSON文件。
    """
    issues = {
        "objective_coverage": [],      # 教学目标覆盖
        "objective_quality": [],       # 教学目标质量
        "hierarchy_depth": [],         # 层级深度
        "hierarchy_jump": [],          # 层级跳跃
        "category_constraint": [],     # 知识分类约束
        "category_coverage": [],       # 知识点分类覆盖
        "name_length": [],             # 名称长度
        "related_coverage": [],        # 关联关系覆盖
        "node_order": [],              # 节点顺序
    }
    
    if not nodes:
        issues["hierarchy_depth"].append("无任何节点")
        return issues
    
    # 统计
    max_level = max(n.level for n in nodes)
    min_level = min(n.level for n in nodes)
    knowledge_nodes = [n for n in nodes if n.node_type == "知识点"]
    classification_nodes = [n for n in nodes if n.node_type == "分类"]
    
    # 1. 层级深度检查：应达到至少6级
    if max_level < 6:
        issues["hierarchy_depth"].append(
            f"层级深度不足：最高为{max_level}级，要求至少达到6级。建议在末级知识点下添加细分节点。"
        )
    
    # 2. 层级跳跃检查：检查是否有断层
    level_set = set(n.level for n in nodes)
    for l in range(min_level, max_level):
        if l not in level_set and l + 1 in level_set:
            issues["hierarchy_jump"].append(
                f"层级断层：缺少{LEVEL_NAMES.get(l, f'{l}级')}节点，但存在{LEVEL_NAMES.get(l+1, f'{l+1}级')}节点"
            )
    
    # 3. 教学目标覆盖：Level 4-7知识点节点应有objective
    for node in knowledge_nodes:
        if node.level >= 4 and not node.objective:
            issues["objective_coverage"].append(
                f"'{node.name}' (level={node.level}) 缺少教学目标，应填写具体可衡量的目标"
            )
    
    # 4. 教学目标质量：应使用行为动词
    behavior_verbs = ["能够", "会", "掌握", "理解", "学会", "说出", "解释", "操作", "分析", "设计", "创建", "判断", "选择"]
    for node in knowledge_nodes:
        if node.objective:
            if not any(v in node.objective for v in behavior_verbs):
                issues["objective_quality"].append(
                    f"'{node.name}' 教学目标未使用行为动词：'{node.objective}'，建议改为'能够...'句式"
                )
            if len(node.objective) < 10:
                issues["objective_quality"].append(
                    f"'{node.name}' 教学目标过短：'{node.objective}'，应描述具体可衡量的学习成果"
                )
    
    # 5. 知识分类约束：分类节点category必须为空
    for node in classification_nodes:
        if node.category:
            issues["category_constraint"].append(
                f"'{node.name}' (分类节点) 的category应为空，实际为'{node.category}'"
            )
    
    # 6. 知识点分类覆盖：知识点节点应有category
    for node in knowledge_nodes:
        if not node.category and node.level >= 4:
            issues["category_coverage"].append(
                f"'{node.name}' (知识点) 缺少知识分类，应填写：事实性/概念性/程序性/元认知"
            )
    
    # 7. 名称长度检查
    for node in nodes:
        if len(node.name) > 30:
            issues["name_length"].append(
                f"'{node.name}' 名称过长({len(node.name)}字)，建议缩短至30字以内"
            )
    
    # 8. 关联关系覆盖
    nodes_without_related = [n for n in knowledge_nodes if not n.related]
    coverage_rate = (len(knowledge_nodes) - len(nodes_without_related)) / len(knowledge_nodes) * 100 if knowledge_nodes else 0
    if coverage_rate < 50:
        issues["related_coverage"].append(
            f"关联关系覆盖率过低：{coverage_rate:.1f}%（要求≥50%）。{len(nodes_without_related)}个知识点无关联关系。"
        )
        # 列出具体节点（最多10个）
        for n in nodes_without_related[:10]:
            issues["related_coverage"].append(f"  - '{n.name}' 缺少关联关系")
    
    # 9. 节点顺序检查（深度优先）
    level_stack = {}
    for i, node in enumerate(nodes):
        # 检查是否有父节点
        expected_parent_level = node.level - 1
        if expected_parent_level >= min_level:
            if expected_parent_level not in level_stack:
                issues["node_order"].append(
                    f"'{node.name}' (位置{i+1}) 出现层级跳跃，缺少level={expected_parent_level}的父节点"
                )
        # 清理栈
        for l in list(level_stack.keys()):
            if l >= node.level:
                del level_stack[l]
        level_stack[node.level] = node.name
    
    return issues

def generate_quality_report(issues: Dict[str, List[str]], nodes: List[KnowledgeNode]) -> str:
    """生成质量校验报告"""
    total_issues = sum(len(v) for v in issues.values())
    passed_categories = sum(1 for v in issues.values() if not v)
    total_categories = len(issues)
    
    # 统计节点信息
    knowledge_count = sum(1 for n in nodes if n.node_type == "知识点")
    classification_count = sum(1 for n in nodes if n.node_type == "分类")
    max_level = max(n.level for n in nodes) if nodes else 0
    
    report = f"""
{'=' * 60}
质量校验报告
{'=' * 60}

节点统计:
  - 总节点数: {len(nodes)} 条
  - 分类节点: {classification_count} 条
  - 知识点节点: {knowledge_count} 条
  - 最高层级: {max_level} 级

校验结果:
  - 通过项: {passed_categories}/{total_categories}
  - 发现问题: {total_issues} 个
"""
    
    # 按类别输出问题
    for category, problems in issues.items():
        status = "✅ 通过" if not problems else f"❌ {len(problems)}个问题"
        report += f"\n【{category}】{status}\n"
        if problems:
            for p in problems:
                report += f"  {p}\n"
    
    # 修正建议
    if total_issues > 0:
        report += f"""
{'=' * 60}
修正建议
{'=' * 60}

Agent应根据以上问题清单逐一修正JSON文件：
1. 使用 Edit 工具修改不符合规范的字段值
2. 补充缺失的节点或字段
3. 调整节点顺序为深度优先
4. 修正完成后重新运行校验

校验命令:
  python scripts/extract_knowledge_graph.py --json {nodes[0].name if nodes else 'nodes'}.json --validate-only
"""
    else:
        report += f"\n{'=' * 60}\n✅ 所有校验项通过，JSON文件质量合格\n{'=' * 60}\n"
    
    return report

# ==================== 统计输出 ====================

def print_statistics(nodes: List[KnowledgeNode], rows: List[Dict], output_path: str, csv_path: str):
    """输出统计信息"""
    level_dist = {}
    type_dist = {}
    for node in nodes:
        level_dist[node.level] = level_dist.get(node.level, 0) + 1
        type_dist[node.node_type] = type_dist.get(node.node_type, 0) + 1
    
    i_count = sum(1 for n in nodes if n.pre_requisites)
    j_count = sum(1 for n in nodes if n.post_requisites)
    k_count = sum(1 for n in nodes if n.related)
    
    # 验证树状结构（B-H列每行只有一个）
    tree_valid = all(sum(1 for c in 'BCDEFGH' if r.get(c)) == 1 for r in rows)
    
    # 验证节点类型与知识分类约束
    type_cat_valid = all(
        (r.get('A') == "分类" and not r.get('M')) or 
        (r.get('A') == "知识点" and (not r.get('M') or r.get('M') in VALID_CATEGORIES))
        for r in rows
    )
    
    print(f"""
{'=' * 60}
抽取结果概要
{'=' * 60}

节点统计:
  - 节点总数: {len(nodes)} 条
  - 节点类型分布: {', '.join([f'{t}: {c}' for t, c in type_dist.items()])}
  - 层级分布: {', '.join([f'{LEVEL_NAMES.get(l, l)}: {c}' for l, c in sorted(level_dist.items())])}

关系统计:
  - 前置关系(I列): {i_count} 条
  - 后置关系(J列): {j_count} 条
  - 关联关系(K列): {k_count} 条

格式验证:
  - 节点类型: {'✅ 通过' if all(r.get('A') in NODE_TYPES for r in rows) else '❌ 失败'}
  - 树状结构(B-H列): {'✅ 通过' if tree_valid else '❌ 失败'}
  - 知识分类约束(M列): {'✅ 通过' if type_cat_valid else '❌ 失败'}

文件输出:
  - Excel路径: {output_path}
  - CSV备份: {csv_path}
""")

# ==================== 主函数 ====================

def main():
    parser = argparse.ArgumentParser(
        description='文档知识图谱结构化抽取',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 从 LLM JSON 生成 Excel（推荐）
  %(prog)s --json nodes.json --template template.xlsx --output output.xlsx
  
  # 仅执行质量校验（不生成Excel）
  %(prog)s --json nodes.json --validate-only
  
  # 提取文档内容供 LLM 处理
  %(prog)s --source document.docx --extract-text --output content.txt
  
  # 仅解析模板
  %(prog)s --template template.xlsx --dry-run
""")
    parser.add_argument('--json', help='LLM 输出的 JSON 文件路径')
    parser.add_argument('--source', help='源文档路径 (docx/pdf)')
    parser.add_argument('--extract-text', action='store_true', help='提取文档文本供 LLM 处理')
    parser.add_argument('--template', help='xlsx 模板路径')
    parser.add_argument('--output', help='输出文件路径')
    parser.add_argument('--dry-run', action='store_true', help='仅解析模板')
    parser.add_argument('--validate-only', action='store_true', help='仅执行质量校验，输出问题清单')
    args = parser.parse_args()
    
    # 模式0: 仅质量校验 (validate-only)
    if args.validate_only:
        if not args.json:
            print("错误: validate-only 模式需要 --json 参数", file=sys.stderr)
            sys.exit(1)
        
        print("执行质量校验...")
        nodes = parse_llm_json(args.json)
        print(f"  已加载 {len(nodes)} 个知识节点")
        
        issues = validate_quality(nodes)
        report = generate_quality_report(issues, nodes)
        print(report)
        
        total_issues = sum(len(v) for v in issues.values())
        if total_issues > 0:
            sys.exit(1)  # 返回非零退出码表示有问题
        return
    if args.source and args.extract_text:
        print("步骤1: 提取文档内容...")
        content = extract_document_content(args.source)
        
        if not content:
            print("错误: 无法提取文档内容", file=sys.stderr)
            sys.exit(1)
        
        print(f"  内容长度: {len(content)} 字符")
        
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f"✅ 文档内容已保存: {args.output}")
            print("\n下一步: 请将文档内容传递给 LLM 进行知识抽取，然后使用 --json 参数生成 Excel")
        else:
            print("\n" + "=" * 60)
            print("文档内容:")
            print("=" * 60)
            print(content)
        
        return
    
    # 模式2: 解析模板 (dry-run)
    if args.dry_run:
        if not args.template:
            print("错误: dry-run 模式需要 --template 参数", file=sys.stderr)
            sys.exit(1)
        
        print("解析模板...")
        template = parse_template(args.template)
        print(f"  模板列数: {template['max_column']}")
        print(f"  Sheet名称: {template['sheet_name']}")
        print(f"  A1规则: {template['a1_rules'][:100]}..." if len(template['a1_rules']) > 100 else f"  A1规则: {template['a1_rules']}")
        print("\n✅ 模板解析完成")
        return
    
    # 模式3: 从 JSON 生成 Excel
    if not args.json:
        print("错误: 需要指定 --json 参数", file=sys.stderr)
        print("提示: 使用 --source --extract-text 提取文档内容，然后交给 LLM 处理", file=sys.stderr)
        parser.print_help()
        sys.exit(1)
    
    if not args.template or not args.output:
        print("错误: 需要 --template 和 --output 参数", file=sys.stderr)
        sys.exit(1)
    
    # 步骤1: 解析模板
    print("步骤1: 解析模板...")
    template = parse_template(args.template)
    print(f"  模板列数: {template['max_column']}")
    
    # 步骤2: 解析 JSON
    print("步骤2: 解析 LLM JSON 输出...")
    nodes = parse_llm_json(args.json)
    print(f"  已加载 {len(nodes)} 个知识节点")
    
    # 步骤3: 生成关系
    print("步骤3: 生成知识点关系...")
    generate_post_relations(nodes)
    
    relation_errors = validate_relations(nodes)
    if relation_errors:
        print(f"  ⚠️ 关系验证警告 ({len(relation_errors)} 条):")
        for err in relation_errors[:3]:
            print(f"    - {err}")
        if len(relation_errors) > 3:
            print(f"    - ... 还有 {len(relation_errors) - 3} 条")
    
    # 步骤4: 构建层级树
    print("步骤4: 构建层级结构...")
    root_nodes = build_knowledge_tree(nodes)
    rows = flatten_nodes(root_nodes)
    
    # 步骤5: 生成输出
    print("步骤5: 生成输出文件...")
    csv_path = args.output.replace('.xlsx', '.csv') if args.output.endswith('.xlsx') else args.output + '.csv'
    
    generate_csv(rows, csv_path)
    if args.output.endswith('.xlsx'):
        generate_excel(rows, args.output, template['a1_rules'])
    
    # 步骤6: 验证
    print("步骤6: 输出验证...")
    errors = validate_rows(rows)
    if errors:
        print(f"  ❌ 发现 {len(errors)} 个错误:")
        for i, err in enumerate(errors[:5], 1):
            print(f"     {i}. {err}")
    else:
        print("  ✅ 验证通过")

    # 步骤7: 表头校验
    print("步骤7: 表头校验...")
    header_errors = validate_headers(args.output)
    if header_errors:
        print(f"  ❌ 发现 {len(header_errors)} 个表头错误:")
        for err in header_errors:
            print(f"     {err}")
        sys.exit(1)
    else:
        print("  ✅ 表头校验通过")

    # 输出统计
    print_statistics(nodes, rows, args.output, csv_path)

if __name__ == "__main__":
    main()